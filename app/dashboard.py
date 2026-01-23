"""FastAPI dashboard for exploring Lowe's clearance items."""

from __future__ import annotations

import json
import os
import sqlite3
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
import time
import re
from typing import Any, Iterable, Literal
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

from fastapi import Depends, FastAPI, HTTPException, Query, Request
from fastapi.responses import JSONResponse, PlainTextResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field, field_validator
from sqlalchemy.orm import Session

from app.logging_config import get_logger
from app.middleware.simple_session import SimpleSessionMiddleware
from app.lowes_stores_wa_or import LOWES_STORES_WA_OR
from app.normalizers import extract_category_name
from app.storage import repo
from app.storage.db import get_engine, init_db, make_session, resolve_database_file
from app.storage.models_sql import Observation
from app.timezone_utils import format_regional_timestamp, format_regional_full_timestamp


LOGGER = get_logger(__name__)
BASE_PATH = Path(__file__).resolve().parent
TEMPLATES_DIR = BASE_PATH / "templates"
STATIC_DIR = BASE_PATH / "static"
FALLBACK_DATABASE_FILE = (BASE_PATH.parent / "orwa_lowes.sqlite").resolve()
DATABASE_FILE = resolve_database_file(fallback=FALLBACK_DATABASE_FILE)
METRICS_SUMMARY_FILE = Path(
    os.getenv("CHEAPSKATER_METRICS_SUMMARY", "logs/metrics_summary.json")
)
ZIP_CURSOR_FILE = Path(os.getenv("CHEAPSKATER_ZIP_CURSOR", "logs/zip_cursor.json"))
HEALTH_MAX_STALE_MINUTES = float(os.getenv("DASHBOARD_HEALTH_MAX_STALE_MINUTES", "120"))

DB_BUSY_TIMEOUT = float(os.getenv("DB_BUSY_TIMEOUT", "30"))
engine = get_engine(str(DATABASE_FILE), busy_timeout=DB_BUSY_TIMEOUT)
init_db(engine)
session_factory = make_session(engine)

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
app = FastAPI(title="CheapSkater Clearance Dashboard")

# Register ingest API router
from app.ingest import router as ingest_router

app.include_router(ingest_router)

# Register auth router for login, registration, and subscriptions
from app.auth.routes import router as auth_router

app.include_router(auth_router)

# Register notification routes for deal alerts
from app.notifications.routes import router as notifications_router

app.include_router(notifications_router)

# Register AI assistant routes (unlisted subsite)
from app.assistant.routes import router as assistant_router

app.include_router(assistant_router)

# Register admin routes for system management
from app.admin.routes import router as admin_router

app.include_router(admin_router)


def _ensure_schema_up_to_date():
    """Self-healing: Ensures database has the 'region' column required for Florida expansion."""
    db_path = DATABASE_FILE
    if not db_path.exists():
        return

    LOGGER.info("Verifying database schema: %s", db_path)
    # Give it a retry if locked
    for i in range(3):
        try:
            conn = sqlite3.connect(str(db_path), timeout=30)
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(stores)")
            columns = [row[1] for row in cursor.fetchall()]

            if "region" not in columns:
                LOGGER.warning(
                    "Schema out of date: 'region' column missing. Running auto-migration..."
                )
                cursor.execute("ALTER TABLE stores ADD COLUMN region TEXT")
                try:
                    cursor.execute("ALTER TABLE observations ADD COLUMN region TEXT")
                except sqlite3.OperationalError:
                    pass
                try:
                    cursor.execute(
                        "ALTER TABLE store_price_history ADD COLUMN region TEXT"
                    )
                except sqlite3.OperationalError:
                    pass
                conn.commit()
                LOGGER.info("Auto-migration successful.")

            # Always run region backfill/repair on startup
            # This fixes cases where stores were added (e.g. via ingest) without a region
            cursor.execute("""
                UPDATE stores 
                SET region = CASE 
                    WHEN state IN ('WA', 'OR') THEN 'WA_OR'
                    WHEN state = 'FL' THEN 'FL'
                    ELSE NULL
                END
                WHERE region IS NULL OR region = ''
            """)
            if cursor.rowcount > 0:
                conn.commit()
                LOGGER.info(f"Repaired regions for {cursor.rowcount} stores.")
            else:
                LOGGER.debug("No region repairs needed.")

            conn.close()
            break
        except sqlite3.OperationalError as e:
            if "locked" in str(e).lower() and i < 2:
                time.sleep(1)
                continue
            LOGGER.error("Auto-migration database locked: %s", e)
            break
        except Exception as e:
            LOGGER.exception("Auto-migration failed: %s", e)
            break


# Run self-healing check
# Run self-healing check
_ensure_schema_up_to_date()


def _ensure_admin_permissions():
    """Self-healing: Ensures primary admin account has correct permissions and subscription."""
    target_email = "93robingattis@gmail.com"

    try:
        from app.auth.models import (
            User,
            Subscription,
            SubscriptionPlan,
            SubscriptionStatus,
        )
        from app.admin.models import Base as AdminBase

        LOGGER.info("Creating admin tables...")
        AdminBase.metadata.create_all(engine, checkfirst=True)
        LOGGER.info("Admin tables verified")

        # Use a fresh session directly
        with session_factory() as session:
            user = session.query(User).filter(User.email == target_email).first()
            if not user:
                create_admin_user(session)
                return

            changes_made = False

            # 1. Ensure Admin Status
            if not user.is_admin:
                user.is_admin = True
                changes_made = True
                LOGGER.info(f"Upgraded {target_email} to Admin.")

            # 2. Ensure Subscription Status
            sub = (
                session.query(Subscription)
                .filter(Subscription.user_id == user.id)
                .first()
            )
            if not sub:
                sub = Subscription(
                    user_id=user.id,
                    plan=SubscriptionPlan.PREMIUM,
                    status=SubscriptionStatus.ACTIVE,
                    # is_active property is derived, not stored
                )
                session.add(sub)
                changes_made = True
                LOGGER.info(f"Created Premium subscription for {target_email}.")
            else:
                if (
                    sub.plan != SubscriptionPlan.PREMIUM
                    or sub.status != SubscriptionStatus.ACTIVE
                ):
                    sub.plan = SubscriptionPlan.PREMIUM
                    sub.status = SubscriptionStatus.ACTIVE
                    changes_made = True
                    LOGGER.info(f"Restored Premium subscription for {target_email}.")

            if changes_made:
                session.commit()
                LOGGER.info(f"Admin permissions enforced for {target_email}")

    except Exception as e:
        LOGGER.exception(f"Failed to enforce admin permissions: {e}")


def create_admin_user(session):
    from app.auth.models import User, Subscription, SubscriptionPlan, SubscriptionStatus

    LOGGER.info("Creating admin user: 93robingattis@gmail.com")
    admin_user = User(
        email="93robingattis@gmail.com",
        hashed_password="$2b$12$LQv3c1yqBWVHxkd0LHAkCOYz6TtxMQJqhN8/LewY5GyYKwq5qDx1",
        is_admin=True,
        is_active=True,
    )
    session.add(admin_user)
    session.flush()

    sub = Subscription(
        user_id=admin_user.id,
        plan=SubscriptionPlan.PREMIUM,
        status=SubscriptionStatus.ACTIVE,
    )
    session.add(sub)
    session.commit()
    LOGGER.info("Admin user created successfully")


_ensure_admin_permissions()


class IngestDeal(BaseModel):
    store_id: str
    store_name: str
    category_url: str | None = None
    category_name: str | None = None
    product_url: str
    title: str
    image_url: str | None = None
    price: float
    was_price: float
    pct_off: float
    found_at: str


class IngestRequest(BaseModel):
    source: str
    deals: list[IngestDeal]


def _extract_sku(url: str) -> str | None:
    if not url:
        return None
    # Common patterns: /pd/name/SKU or /product/name/SKU
    # specific lowes pattern often ends in digits
    match = re.search(r"/(\d{4,20})(?:$|[?#])", url)
    if match:
        return match.group(1)
    return None


def _extract_category_name(url: str | None) -> str:
    return extract_category_name(url)


def _sanitize_image_url(value: str | None) -> str | None:
    if not isinstance(value, str):
        return None
    cleaned = value.strip()
    if not cleaned:
        return None
    lowered = cleaned.lower()
    if lowered.startswith("data:"):
        return None
    if lowered.endswith(".svg"):
        return None
    # Filter out badge/icon images including clearance tags
    if any(token in lowered for token in ("badge", "sprite", "icon", "clearance")):
        return None
    return cleaned


SESSION_SECRET = os.getenv("CHEAPSKATER_SESSION_SECRET", "cheapskater-session-secret")
SESSION_MAX_AGE = int(os.getenv("CHEAPSKATER_SESSION_MAX_AGE", str(60 * 60 * 24 * 30)))
app.add_middleware(
    SimpleSessionMiddleware, secret_key=SESSION_SECRET, max_age=SESSION_MAX_AGE
)

# Add user activity tracking middleware
from app.admin.middleware import UserActivityMiddleware

app.add_middleware(UserActivityMiddleware, database_file=str(DATABASE_FILE))
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

_STATS_CACHE: dict[str, Any] | None = None
_STATS_CACHE_TS: float | None = None


def _read_json(path: Path) -> dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        return {}
    except Exception:
        return {}


@app.exception_handler(404)
async def not_found_handler(request: Request, exc: HTTPException) -> PlainTextResponse:
    if STATIC_DIR.exists():
        detail = exc.detail if isinstance(exc.detail, str) else "Not Found"
        return PlainTextResponse(detail, status_code=404)
    return PlainTextResponse(
        "CheapSkater Dashboard UI unavailable.\n"
        "Use: /api/clearance, /api/stats, /healthz",
        status_code=404,
    )


Scope = Literal["all", "new"]
STATE_OPTIONS = ["ALL", "WA", "OR", "FL"]
REGION_OPTIONS = [
    ("ALL", "All Regions"),
    ("WA_OR", "Washington & Oregon"),
    ("FL", "Florida"),
]
SORT_OPTIONS = [
    ("newest", "Newest first"),
    ("alpha_asc", "Title A → Z"),
    ("alpha_desc", "Title Z → A"),
    ("price_low", "Price: Low → High"),
    ("price_high", "Price: High → Low"),
]
DEFAULT_CATEGORY_OPTIONS: list[str] = []
INITIAL_GROUP_BATCH = 30
_STORE_HOURS_SUFFIX = re.compile(r"\s+\d{1,2}\s*(?:AM|PM)$", re.I)
_STOCK_COUNT_PATTERN = re.compile(
    r"(?:only|about)?\s*(\d+)\s*(?:left|available|in\s*stock|qty|quantity)", re.I
)
_SAVED_DEALS_SESSION_KEY = "saved_deals"
_STORE_STATUS_TOKEN_PATTERN = re.compile(r"\s+(?:closed|open)\b", re.I)
_STORE_HOURS_RANGE_PATTERN = re.compile(
    r"\s+\d{1,2}\s*(?:AM|PM)\s*(?:[-–—]\s*\d{1,2}\s*(?:AM|PM))?",
    re.I,
)
_STORE_HOURS_VERB_PATTERN = re.compile(
    r"\s+(?:opens|closes)\s+\d{1,2}\s*(?:AM|PM)", re.I
)


class SaveDealPayload(BaseModel):
    store_number: str = Field(..., description="Lowe's store number.")
    sku: str = Field(..., description="Product SKU.")
    quantity: int | None = Field(
        default=1, ge=1, description="Quantity to add to the saved cart."
    )


def _normalize_store_number(value: str | None) -> str:
    """Return a zero-padded store number for lookups."""

    text = (value or "").strip()
    if text.isdigit() and len(text) < 4:
        return text.zfill(4)
    return text


def _canonical_store_details(store_number: str | None) -> dict[str, str] | None:
    """Look up canonical store metadata from the WA/OR registry."""

    normalized = _normalize_store_number(store_number)
    if not normalized:
        return None
    return LOWES_STORES_WA_OR.get(normalized)


def _strip_store_status_text(value: str | None) -> str | None:
    if not value:
        return None
    text = value.strip()
    if not text:
        return None
    text = _STORE_STATUS_TOKEN_PATTERN.sub("", text)
    text = _STORE_HOURS_VERB_PATTERN.sub("", text)
    text = _STORE_HOURS_RANGE_PATTERN.sub("", text)
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip() or None


def normalize_store_label(store_number: str | None, store_name_raw: str | None) -> str:
    """Resolve a user-facing store label, preferring the canonical address."""

    details = _canonical_store_details(store_number)
    if details and details.get("address"):
        return details["address"]
    cleaned = _strip_store_status_text(store_name_raw)
    if cleaned:
        return cleaned
    normalized_number = _normalize_store_number(store_number)
    if normalized_number:
        return f"Lowe's #{normalized_number}"
    return "Lowe's Store"


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _coerce_datetime(value: object | None) -> datetime | None:
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)
    return None


def _relative_days(value: object | None) -> int | None:
    dt_value = _coerce_datetime(value)
    if not dt_value:
        return None
    delta = _now_utc() - dt_value
    return max(int(delta.total_seconds() // 86400), 0)


def _time_delta_for(value: str | None) -> timedelta | None:
    for key, _, delta in TIME_FILTER_OPTIONS:
        if key == value:
            return delta
    return None


def _sanitize_percent(value: float | None) -> float | None:
    if value is None:
        return None
    return max(0.0, min(100.0, float(value)))


def _sanitize_stock(value: float | None) -> int | None:
    if value is None:
        return None
    return max(0, int(value))


def _as_float(value: float | str | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = value.strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _as_int(value: int | str | None) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = value.strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _estimate_stock_units(availability: str | None) -> int | None:
    if not availability:
        return None
    normalized = availability.strip().lower()
    if not normalized:
        return None
    match = _STOCK_COUNT_PATTERN.search(normalized)
    if match:
        try:
            return max(int(match.group(1)), 0)
        except (TypeError, ValueError):
            return None
    if "out of stock" in normalized:
        return 0
    if "limited" in normalized:
        return 1
    if "only" in normalized:
        return 1
    return None


def _format_stock_status(
    stock_estimate: int | None, availability: str | None
) -> str | None:
    if stock_estimate is None:
        return None
    if stock_estimate == 0:
        return "0 in stock"
    return f"{stock_estimate} in stock"


def _prepare_listing(listing: dict[str, Any]) -> dict[str, Any]:
    enriched = dict(listing)
    store_number = _normalize_store_number(
        enriched.get("store_id") or enriched.get("store_number")
    )
    canonical_store = _canonical_store_details(store_number)
    if canonical_store:
        enriched.setdefault("store_name", canonical_store.get("name"))
        enriched.setdefault("store_city", canonical_store.get("city"))
        enriched.setdefault("store_state", canonical_store.get("state"))
        enriched.setdefault("store_zip", canonical_store.get("zip"))
    if not enriched.get("store_state"):
        enriched["store_state"] = _state_from_zip(enriched.get("store_zip"))
    enriched["store_label"] = _format_store_label(enriched)
    enriched["store_tooltip"] = _format_store_tooltip(enriched)
    enriched["store_product_url"] = _store_specific_url(
        enriched.get("product_url"), enriched.get("store_id")
    )
    added_ts = enriched.get("first_seen") or enriched.get("price_started_at")
    enriched["days_since_added"] = _relative_days(added_ts)
    stock_estimate = _estimate_stock_units(enriched.get("availability"))
    enriched["stock_estimate"] = stock_estimate
    enriched["stock_label"] = _format_stock_status(
        stock_estimate, enriched.get("availability")
    )
    return enriched


def _prepare_listings(listings: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    return [_prepare_listing(listing) for listing in listings]


def _apply_filters(
    listings: Iterable[dict[str, Any]],
    *,
    filters: dict[str, Any],
) -> list[dict[str, Any]]:
    cutoff = filters.get("time_cutoff")
    discount_min = filters.get("discount_min")
    stock_min = filters.get("stock_min")
    stock_max = filters.get("stock_max")

    result: list[dict[str, Any]] = []
    for listing in listings:
        added_at = listing.get("first_seen") or listing.get("price_started_at")
        added_dt = _coerce_datetime(added_at)
        if cutoff and (added_dt is None or added_dt < cutoff):
            continue
        pct_off = listing.get("pct_off")
        if discount_min is not None:
            if pct_off is None or pct_off < discount_min:
                continue
        stock_estimate = listing.get("stock_estimate")
        if stock_min is not None:
            if stock_estimate is None or stock_estimate < stock_min:
                continue
        if stock_max is not None:
            if stock_estimate is None or stock_estimate > stock_max:
                continue
        result.append(listing)
    return result


def _normalize_filters(
    *,
    time_window: str | None,
    discount_filter: str | None,
    discount_min: str | float | None,
    discount_max: str | float | None,
    stock_filter: str | None,
    stock_min: str | int | None,
    stock_max: str | int | None,
    sort_order: str | None,
) -> dict[str, Any]:
    normalized_time = (
        time_window
        if any(opt[0] == time_window for opt in TIME_FILTER_OPTIONS)
        else "all"
    )
    delta = _time_delta_for(normalized_time)
    cutoff = None
    if delta:
        cutoff = _now_utc() - delta

    discount_choice = discount_filter or "all"
    preset_discounts = {
        value: min_value
        for value, _, min_value, _ in DISCOUNT_FILTER_OPTIONS
        if min_value is not None and value not in {"all", "custom"}
    }
    min_pct = None
    max_pct = None
    if discount_choice in preset_discounts:
        min_pct = preset_discounts[discount_choice]
    elif discount_choice == "custom":
        parsed_min = _sanitize_percent(_as_float(discount_min))
        parsed_max = _sanitize_percent(_as_float(discount_max))
        min_pct = parsed_min
        max_pct = parsed_max
        if min_pct is not None and max_pct is not None and max_pct < min_pct:
            min_pct, max_pct = max_pct, min_pct
    else:
        discount_choice = "all"

    stock_choice = stock_filter or "all"
    preset_stock = {
        value: min_value
        for value, _, min_value, _ in STOCK_FILTER_OPTIONS
        if min_value is not None and value not in {"all", "custom"}
    }
    stock_floor = None
    stock_ceiling = None
    if stock_choice in preset_stock:
        stock_floor = preset_stock[stock_choice]
    elif stock_choice == "custom":
        parsed_min = _as_int(stock_min)
        parsed_max = _as_int(stock_max)
        stock_floor = _sanitize_stock(parsed_min)
        stock_ceiling = _sanitize_stock(parsed_max)
        if (
            stock_floor is not None
            and stock_ceiling is not None
            and stock_ceiling < stock_floor
        ):
            stock_floor, stock_ceiling = stock_ceiling, stock_floor
    else:
        stock_choice = "all"

    sort_choice = (
        sort_order if sort_order in {value for value, _ in SORT_OPTIONS} else "newest"
    )

    return {
        "time_window": normalized_time,
        "time_cutoff": cutoff,
        "discount_choice": discount_choice,
        "discount_min_pct": min_pct,
        "discount_max_pct": max_pct,
        "discount_min": (min_pct / 100.0) if min_pct is not None else None,
        "discount_max": (max_pct / 100.0) if max_pct is not None else None,
        "stock_choice": stock_choice,
        "stock_min": stock_floor,
        "stock_max": stock_ceiling,
        "sort_choice": sort_choice,
    }


def _spread(min_value: float | None, max_value: float | None) -> float | None:
    if min_value is None or max_value is None:
        return None
    delta = max_value - min_value
    return delta if delta > 0 else 0.0


TIME_FILTER_OPTIONS = [
    ("all", "All Dates", None),
    ("1h", "Last 1 hour", timedelta(hours=1)),
    ("12h", "Last 12 hours", timedelta(hours=12)),
    ("24h", "Last 24 hours", timedelta(days=1)),
    ("3d", "Last 3 days", timedelta(days=3)),
    ("1w", "Last 1 week", timedelta(weeks=1)),
    ("1m", "Last 1 month", timedelta(days=30)),
]
DISCOUNT_FILTER_OPTIONS = [
    ("all", "All Discounts", None, None),
    ("50", "50%+", 50.0, None),
    ("60", "60%+", 60.0, None),
    ("75", "75%+", 75.0, None),
    ("90", "90%+", 90.0, None),
    ("custom", "Custom Range…", None, None),
]
STOCK_FILTER_OPTIONS = [
    ("all", "Any Stock", None, None),
    ("1", "1+ Stock", 1, None),
    ("2", "2+ Stock", 2, None),
    ("3", "3+ Stock", 3, None),
    ("4", "4+ Stock", 4, None),
    ("5", "5+ Stock", 5, None),
    ("custom", "Custom Range…", None, None),
]


def _sort_groups(
    groups: list[dict[str, Any]], sort_choice: str | None
) -> list[dict[str, Any]]:
    """Return *groups* ordered according to the requested sort."""

    choice = sort_choice or "newest"

    def _timestamp_key(group: dict[str, Any]) -> datetime:
        ts = group.get("last_seen") or group.get("added_at")
        dt_value = _coerce_datetime(ts)
        if dt_value:
            return dt_value
        return datetime.fromtimestamp(0, timezone.utc)

    def _alpha_key(group: dict[str, Any]) -> tuple[str, str]:
        title = (group.get("title") or "").strip().lower()
        sku = (group.get("sku") or "").strip().lower()
        return (title, sku)

    def _price_key(group: dict[str, Any]) -> tuple[int, float]:
        price = group.get("min_price")
        if price is None:
            return (1, float("inf"))
        return (0, float(price))

    def _price_desc_key(group: dict[str, Any]) -> tuple[int, float]:
        price = group.get("min_price")
        if price is None:
            return (1, 0.0)
        return (0, -float(price))

    if choice == "alpha_asc":
        groups.sort(key=_alpha_key)
    elif choice == "alpha_desc":
        groups.sort(key=_alpha_key, reverse=True)
    elif choice == "price_low":
        groups.sort(key=_price_key)
    elif choice == "price_high":
        groups.sort(key=_price_desc_key)
    else:
        groups.sort(key=_timestamp_key, reverse=True)
    return groups


def _store_specific_url(url: str | None, store_id: str | None) -> str | None:
    if not url:
        return None
    if not store_id:
        return url

    parsed = urlparse(url)
    params = parse_qsl(parsed.query, keep_blank_values=True)
    has_store = any(key.lower() == "storenumber" for key, _ in params)
    if not has_store:
        params.append(("storeNumber", store_id))
    new_query = urlencode(params, doseq=True)
    return urlunparse(parsed._replace(query=new_query))


def _clean_store_name(value: str | None) -> str | None:
    if not value:
        return None
    text = value.strip()
    if not text:
        return None
    return _STORE_HOURS_SUFFIX.sub("", text)


def _format_store_label(listing: dict[str, Any]) -> str:
    details = _canonical_store_details(listing.get("store_id"))
    if details and details.get("address"):
        return details["address"]

    name = _clean_store_name(listing.get("store_name"))
    city = (listing.get("store_city") or "").strip()
    state = (listing.get("store_state") or "").strip()
    store_id = _normalize_store_number(listing.get("store_id"))

    label = name or ""
    if not label and city:
        label = f"{city} Lowe's"
    if not label and store_id:
        label = f"Lowe's #{store_id}"
    if not label:
        label = "Lowe's Store"
    if city and city.lower() not in label.lower():
        label = f"{label} — {city}"
    elif state and state not in label:
        label = f"{label} ({state})"
    return label


def _format_store_tooltip(listing: dict[str, Any]) -> str:
    details = _canonical_store_details(listing.get("store_id"))
    if details:
        store_id = _normalize_store_number(listing.get("store_id"))
        parts = [details.get("name")]
        if store_id:
            parts.append(f"Store #{store_id}")
        if details.get("address"):
            parts.append(details["address"])
        return " • ".join(part for part in parts if part)

    city = (listing.get("store_city") or "").strip()
    state = (listing.get("store_state") or "").strip()
    zip_code = (listing.get("store_zip") or "").strip()
    store_id = _normalize_store_number(listing.get("store_id"))
    parts = [part for part in [city, state, zip_code] if part]
    details = ", ".join(parts)
    if store_id:
        details = f"{details} • Store #{store_id}" if details else f"Store #{store_id}"
    return details or "Store location"


def get_session() -> Iterable[Session]:
    """Dependency that yields a SQLAlchemy session."""

    with session_factory() as session:
        yield session


def _cache_stats(session: Session) -> dict[str, Any]:
    global _STATS_CACHE, _STATS_CACHE_TS
    now = time.monotonic()
    if _STATS_CACHE and _STATS_CACHE_TS and (now - _STATS_CACHE_TS) < 60:
        return _STATS_CACHE
    total_items = repo.count_observations(session)
    quarantine_total = repo.count_quarantine(session)
    last_scrape = repo.get_latest_timestamp(session)
    payload = {
        "total_items": total_items,
        "quarantine_count": quarantine_total,
        "last_scrape": last_scrape.isoformat() if last_scrape else None,
    }
    _STATS_CACHE, _STATS_CACHE_TS = payload, now
    return payload


def _serialize_observation(listing: dict[str, Any]) -> dict[str, Any]:
    """Convert a store listing mapping into JSON-serialisable data."""

    return _serialize_listing(listing)


def _format_timestamp(value: Any, *, show_time: bool = True) -> str | None:
    if not value:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        try:
            value = datetime.fromisoformat(text)
        except ValueError:
            return text
    if not isinstance(value, datetime):
        return str(value)
    target = value.astimezone(timezone.utc)
    dst = target.dst()
    pacific_offset_hours = -7 if (dst and dst.total_seconds() != 0) else -8
    pacific = target.astimezone(
        timezone(timedelta(hours=pacific_offset_hours), name="PT")
    )
    fmt = "%b %d %I:%M %p" if show_time else "%b %d"
    return pacific.strftime(fmt) + " PT"


def _format_currency(value: float | None) -> str | None:
    if value is None:
        return None
    return f"${value:,.2f}"


def _normalize_state(value: str | None) -> str | None:
    if not value:
        return None
    upper = value.upper()
    return upper if upper in {"WA", "OR"} else None


def _listing_state(listing: dict[str, Any]) -> str | None:
    """Best-effort state inference for a listing."""

    direct = listing.get("store_state")
    if isinstance(direct, str) and direct.strip():
        return direct.strip().upper()

    candidate_zip = listing.get("store_zip") or listing.get("zip")
    inferred = _state_from_zip(candidate_zip)
    if inferred:
        return inferred

    store_id = _normalize_store_number(listing.get("store_id"))
    details = _canonical_store_details(store_id)
    if details and details.get("state"):
        return details["state"]
    return None


def _filter_by_state(
    listings: Iterable[dict[str, Any]], state: str | None
) -> list[dict[str, Any]]:
    if not state:
        return list(listings)
    target = state.upper()
    return [listing for listing in listings if _listing_state(listing) == target]


def _state_from_zip(zip_code: str | None) -> str | None:
    if not zip_code:
        return None
    digits = "".join(ch for ch in zip_code if ch.isdigit())
    if len(digits) < 3:
        return None
    prefix = int(digits[:3])
    if 970 <= prefix <= 979:
        return "OR"
    if 980 <= prefix <= 994:
        return "WA"
    return None


def _calculate_discount_pct(price: float | None, price_was: float | None) -> float:
    """Calculate discount percentage from price and price_was.

    Returns a value 0-100. Returns 0 if prices are invalid or no discount.
    """
    if price is None or price_was is None:
        return 0.0
    if price_was <= 0 or price <= 0:
        return 0.0
    if price >= price_was:
        return 0.0
    return round(((price_was - price) / price_was) * 100.0, 2)


def _serialize_listing(listing: dict[str, Any]) -> dict[str, Any]:
    def _ts(value: Any) -> str | None:
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, str):
            text = value.strip()
            return text or None
        return value

    # Always recalculate pct_off from prices for accuracy
    # This fixes data where pct_off was incorrectly stored as 0 or wrong values
    price = listing.get("price")
    price_was = listing.get("price_was")

    if price and price_was and price_was > price:
        effective_pct_off = _calculate_discount_pct(price, price_was)
    else:
        effective_pct_off = listing.get("pct_off") or 0

    payload = {
        "history_id": listing.get("history_id"),
        "retailer": listing.get("retailer"),
        "store_id": listing.get("store_id"),
        "store_name": listing.get("store_name"),
        "store_city": listing.get("store_city"),
        "store_state": listing.get("store_state"),
        "store_zip": listing.get("store_zip"),
        "sku": listing.get("sku"),
        "title": listing.get("title"),
        "category": listing.get("category"),
        "price": listing.get("price"),
        "price_was": listing.get("price_was"),
        "pct_off": effective_pct_off,
        "availability": listing.get("availability"),
        "product_url": listing.get("product_url"),
        "store_product_url": _store_specific_url(
            listing.get("product_url"), listing.get("store_id")
        ),
        "image_url": _sanitize_image_url(listing.get("image_url")),
        "clearance": listing.get("clearance"),
        "first_seen": _ts(listing.get("first_seen")),
        "price_started_at": _ts(listing.get("price_started_at")),
        "updated_at": _ts(listing.get("updated_at")),
        "prev_price": listing.get("prev_price"),
        "prev_price_was": listing.get("prev_price_was"),
        "prev_pct_off": listing.get("prev_pct_off"),
        "prev_updated_at": _ts(listing.get("prev_updated_at")),
        "prev_clearance": listing.get("prev_clearance"),
        "store_label": _format_store_label(listing),
        "store_tooltip": _format_store_tooltip(listing),
        "store_product_url": listing.get("store_product_url")
        or _store_specific_url(listing.get("product_url"), listing.get("store_id")),
        "stock_estimate": listing.get("stock_estimate"),
        "stock_label": listing.get("stock_label"),
        "days_since_added": _relative_days(
            listing.get("first_seen") or listing.get("price_started_at")
        ),
    }
    canonical = _canonical_store_details(listing.get("store_id"))
    if canonical:
        if canonical.get("address"):
            payload["store_address"] = canonical["address"]
        if canonical.get("city") and not (payload.get("store_city") or "").strip():
            payload["store_city"] = canonical["city"]
        if canonical.get("state") and not (payload.get("store_state") or "").strip():
            payload["store_state"] = canonical["state"]
        if canonical.get("zip") and not (payload.get("store_zip") or "").strip():
            payload["store_zip"] = canonical["zip"]
        if canonical.get("name") and not (payload.get("store_name") or "").strip():
            payload["store_name"] = canonical["name"]
    return payload


def _group_listings(listings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for listing in listings:
        sku = listing.get("sku") or listing.get("history_id")
        if not sku:
            continue
        image_url = _sanitize_image_url(listing.get("image_url"))
        bucket = grouped.setdefault(
            sku,
            {
                "sku": sku,
                "title": listing.get("title"),
                "category": listing.get("category"),
                "image_url": image_url,
                "stores": [],
                "added_at": listing.get("first_seen")
                or listing.get("price_started_at"),
                "last_seen": listing.get("updated_at")
                or listing.get("price_started_at"),
                "days_since_added": listing.get("days_since_added"),
            },
        )
        if not bucket["title"] and listing.get("title"):
            bucket["title"] = listing.get("title")
        if not bucket["category"] and listing.get("category"):
            bucket["category"] = listing.get("category")
        if not bucket["image_url"] and image_url:
            bucket["image_url"] = image_url
        listing_added = listing.get("first_seen") or listing.get("price_started_at")
        if listing_added and (
            not bucket["added_at"]
            or (
                isinstance(listing_added, datetime)
                and isinstance(bucket["added_at"], datetime)
                and listing_added < bucket["added_at"]
            )
        ):
            bucket["added_at"] = listing_added
        listing_seen = listing.get("updated_at") or listing.get("price_started_at")
        if listing_seen and (
            not bucket["last_seen"]
            or (
                isinstance(listing_seen, datetime)
                and isinstance(bucket["last_seen"], datetime)
                and listing_seen > bucket["last_seen"]
            )
        ):
            bucket["last_seen"] = listing_seen
        bucket["stores"].append(listing)

    grouped_list: list[dict[str, Any]] = []
    for bucket in grouped.values():
        stores = bucket["stores"]
        stores.sort(
            key=lambda row: (
                row.get("price") if row.get("price") is not None else float("inf"),
                row.get("pct_off") if row.get("pct_off") is not None else 0,
            )
        )
        prices = [row.get("price") for row in stores if row.get("price") is not None]
        # Always recalculate discounts from price/price_was for accuracy
        # This fixes data where pct_off was incorrectly stored as 0 or wrong values
        discounts = []
        for row in stores:
            price = row.get("price")
            price_was = row.get("price_was")
            if price and price_was and price_was > price:
                calc_pct = _calculate_discount_pct(price, price_was)
                if calc_pct > 0:
                    discounts.append(calc_pct)

        # For the headline max discount, use the best store (lowest price) with its price_was
        # This ensures the badge shows the actual best deal percentage
        if stores:
            best_store = min(stores, key=lambda r: r.get("price") or float("inf"))
            best_price = best_store.get("price")
            best_price_was = best_store.get("price_was")
            if best_price and best_price_was and best_price_was > best_price:
                best_discount = _calculate_discount_pct(best_price, best_price_was)
                if best_discount > 0 and (
                    not discounts or best_discount > max(discounts)
                ):
                    discounts.append(best_discount)

        savings = []
        for row in stores:
            price = row.get("price")
            price_was = row.get("price_was")
            if price is None or price_was is None:
                continue
            savings.append(max(price_was - price, 0))
        stock_values = [
            row.get("stock_estimate")
            for row in stores
            if row.get("stock_estimate") is not None
        ]
        bucket["min_price"] = min(prices) if prices else None
        bucket["max_price"] = max(prices) if prices else None
        bucket["min_pct_off"] = min(discounts) if discounts else None
        bucket["max_pct_off"] = max(discounts) if discounts else None
        bucket["price_spread"] = _spread(bucket["min_price"], bucket["max_price"])
        bucket["discount_spread"] = _spread(
            bucket["min_pct_off"], bucket["max_pct_off"]
        )
        bucket["min_stock_estimate"] = min(stock_values) if stock_values else None
        bucket["max_stock_estimate"] = max(stock_values) if stock_values else None
        bucket["min_savings"] = min(savings) if savings else None
        bucket["max_savings"] = max(savings) if savings else None
        bucket["locations"] = len(stores)
        if stores:
            best_store = stores[0]
            bucket["best_product_url"] = best_store.get(
                "store_product_url"
            ) or best_store.get("product_url")
        else:
            bucket["best_product_url"] = None
        bucket["days_since_added"] = bucket.get("days_since_added") or _relative_days(
            bucket.get("added_at")
        )
        grouped_list.append(bucket)

    grouped_list.sort(
        key=lambda row: (
            row.get("min_price") if row.get("min_price") is not None else float("inf"),
            row.get("sku"),
        )
    )
    return grouped_list


def _serialize_group(group: dict[str, Any]) -> dict[str, Any]:
    stores = group.get("stores", [])

    # Get price_was from the store with min_price (stores are already sorted by price)
    # This ensures we show the correct "was" price for the best deal
    best_store = stores[0] if stores else None
    best_price_was = best_store.get("price_was") if best_store else None

    # Still collect all price_was values for range calculation
    price_was_values = [
        store.get("price_was") for store in stores if store.get("price_was") is not None
    ]

    payload = {
        "sku": group.get("sku"),
        "title": group.get("title"),
        "category": group.get("category"),
        "image_url": group.get("image_url"),
        "min_price": group.get("min_price"),
        "max_price": group.get("max_price"),
        # Use the price_was from the best deal store, not the global minimum
        "min_price_was": best_price_was,
        "max_price_was": max(price_was_values) if price_was_values else None,
        "min_pct_off": group.get("min_pct_off"),
        "max_pct_off": group.get("max_pct_off"),
        "price_spread": group.get("price_spread"),
        "discount_spread": group.get("discount_spread"),
        "min_stock_estimate": group.get("min_stock_estimate"),
        "max_stock_estimate": group.get("max_stock_estimate"),
        "min_savings": group.get("min_savings"),
        "max_savings": group.get("max_savings"),
        "locations": group.get("locations"),
        "added_at": group.get("added_at").isoformat()
        if isinstance(group.get("added_at"), datetime)
        else group.get("added_at"),
        "last_seen": group.get("last_seen").isoformat()
        if isinstance(group.get("last_seen"), datetime)
        else group.get("last_seen"),
        "days_since_added": group.get("days_since_added")
        or _relative_days(group.get("added_at")),
        "best_product_url": group.get("best_product_url"),
        "stores": [_serialize_listing(store) for store in stores],
    }
    return payload


def _coerce_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _coerce_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _datetime_to_iso(value: object | None) -> str | None:
    if isinstance(value, datetime):
        target = value
        if target.tzinfo is None:
            target = target.replace(tzinfo=timezone.utc)
        else:
            target = target.astimezone(timezone.utc)
        return target.isoformat()
    return None


def _build_cheapskater_deal(listing: dict[str, Any]) -> dict[str, Any]:
    store_number_raw = (listing.get("store_id") or "").strip()
    store_number = _normalize_store_number(store_number_raw) or store_number_raw
    sku = (listing.get("sku") or "").strip()
    deal_id = f"{store_number}:{sku}" if store_number or sku else sku
    price_value = _coerce_float(listing.get("price"))
    price_was_value = listing.get("price_was")
    price_was = _coerce_float(price_was_value) if price_was_value is not None else None
    pct_off_raw = _coerce_float(listing.get("pct_off"))
    pct_off_value = None
    if pct_off_raw is not None:
        pct_off_value = pct_off_raw if pct_off_raw > 1 else pct_off_raw * 100
    stock_estimate = listing.get("stock_estimate")
    stock_value = _coerce_int(stock_estimate)
    store_name_raw = listing.get("store_name") or ""
    last_updated = listing.get("updated_at") or listing.get("price_started_at")
    last_updated_label = _format_timestamp(last_updated)
    return {
        "deal_id": deal_id,
        "store_number": store_number,
        "store_name_raw": store_name_raw,
        "store_label": normalize_store_label(store_number, store_name_raw),
        "store_city": listing.get("store_city"),
        "store_state": listing.get("store_state"),
        "store_zip": listing.get("store_zip"),
        "state": listing.get("store_state"),
        "zip": listing.get("store_zip"),
        "product_name": listing.get("title"),
        "sku": sku,
        "category": listing.get("category"),
        "price": price_value,
        "price_display": _format_currency(price_value),
        "price_was": price_was,
        "price_was_display": _format_currency(price_was)
        if price_was is not None
        else None,
        "pct_off": pct_off_value,
        "pct_off_label": f"{pct_off_value:.0f}% off"
        if pct_off_value is not None
        else None,
        "availability": listing.get("availability"),
        "stock": stock_value if stock_value is not None else 0,
        "stock_estimate": stock_estimate,
        "stock_label": listing.get("stock_label"),
        "last_updated": _datetime_to_iso(last_updated),
        "last_updated_label": last_updated_label,
        "days_since_added": listing.get("days_since_added"),
        "product_url": listing.get("product_url"),
        "image_url": listing.get("image_url"),
        "price_started_at": listing.get("price_started_at"),
    }


def _lookup_listing_for_save(
    session: Session,
    store_number: str,
    sku: str,
) -> dict[str, Any] | None:
    candidates: list[str] = []
    raw = (store_number or "").strip()
    if raw:
        candidates.append(raw)
    normalized = _normalize_store_number(raw)
    if normalized and normalized not in candidates:
        candidates.append(normalized)
    trimmed = normalized.lstrip("0") if normalized else raw.lstrip("0")
    if trimmed and trimmed not in candidates:
        candidates.append(trimmed)
    for candidate in candidates:
        if not candidate:
            continue
        listing = repo.get_listing_for_store_and_sku(
            session, store_id=candidate, sku=sku
        )
        if listing:
            return listing
    return None


def _get_saved_deals(request: Request) -> dict[str, dict[str, Any]]:
    stored = request.session.get(_SAVED_DEALS_SESSION_KEY)
    if isinstance(stored, dict):
        return dict(stored)
    request.session[_SAVED_DEALS_SESSION_KEY] = {}
    return {}


def _persist_saved_deals(request: Request, saved: dict[str, dict[str, Any]]) -> None:
    request.session[_SAVED_DEALS_SESSION_KEY] = saved


def _saved_entry_from_deal(deal: dict[str, Any], quantity: int) -> dict[str, Any]:
    return {
        "deal_id": deal.get("deal_id"),
        "store_number": deal.get("store_number"),
        "store_label": deal.get("store_label"),
        "product_name": deal.get("product_name"),
        "sku": deal.get("sku"),
        "price": deal.get("price"),
        "stock": deal.get("stock"),
        "quantity": quantity,
        "last_updated": deal.get("last_updated"),
        "product_url": deal.get("product_url"),
    }


def _cart_totals(saved: dict[str, dict[str, Any]]) -> tuple[int, float]:
    total_items = 0
    total_price = 0.0
    for entry in saved.values():
        qty = int(entry.get("quantity", 0) or 0)
        price = _coerce_float(entry.get("price"))
        if qty < 0:
            qty = 0
        total_items += qty
        total_price += price * qty
    return total_items, round(total_price, 2)


def _deal_line_total(entry: dict[str, Any]) -> float:
    qty = int(entry.get("quantity", 0) or 0)
    if qty < 0:
        qty = 0
    price = _coerce_float(entry.get("price"))
    return round(price * qty, 2)


def _serialize_saved_deal(entry: dict[str, Any]) -> dict[str, Any]:
    payload = dict(entry)
    payload["line_total"] = _deal_line_total(entry)
    return payload


def _group_saved_deals(
    saved: dict[str, dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in saved.values():
        store_number = entry.get("store_number") or "unknown"
        grouped[store_number].append(entry)
    return dict(grouped)


def _get_user_subscription_info(
    session: Session, user: Any | None
) -> tuple[bool, str | None]:
    """
    Get user subscription information.

    Returns:
        tuple[bool, str | None]: (is_paid, plan_name)
        - is_paid: True if user has active paid subscription
        - plan_name: 'free', 'pro', 'premium', or None if no user
    """
    try:
        if not user:
            return False, None
        from app.auth.models import Subscription, SubscriptionPlan

        sub = (
            session.query(Subscription).filter(Subscription.user_id == user.id).first()
        )
        if not sub:
            LOGGER.debug(f"User {user.email} has no subscription record")
            return False, "free"

        is_paid = bool(
            sub.plan != SubscriptionPlan.FREE and sub.is_active_subscription()
        )
        plan_name = sub.plan.value
        LOGGER.debug(
            f"User {user.email}: plan={plan_name}, status={sub.status.value}, is_active={sub.is_active_subscription()}, is_paid={is_paid}"
        )
        return is_paid, plan_name
    except Exception as e:
        LOGGER.error(f"Error checking paid user status: {e}", exc_info=True)
        return False, None


def _is_paid_user(session: Session, user: Any | None) -> bool:
    """Legacy function for backwards compatibility. Use _get_user_subscription_info instead."""
    is_paid, _ = _get_user_subscription_info(session, user)
    return is_paid


def _select_items(
    session: Session,
    *,
    scope: Scope,
    state: str | None,
    category: str | None,
    region: str | None = None,
    user=None,
    is_paid: bool = False,
) -> list[dict[str, Any]]:
    """
    Get clearance items based on scope and user subscription tier.
    Free tier: only sees deals 5+ days old
    Pro/Paid tier: sees all deals
    """
    if scope == "new":
        # New items are always accessible, OR maybe we want to restrict "new today" to pros?
        # For now, let's assume the restriction applies generally to the main feed.
        # Use existing logic for new items.
        return repo.get_new_clearance_today(
            session, state=state, category=category, region=region
        )

    if is_paid:
        return repo.get_clearance_items(
            session, state=state, category=category, region=region
        )

    # Free tier sees older deals
    return repo.get_older_clearance_items(
        session, state=state, category=category, region=region, min_days_old=5
    )


def _collect_categories(
    session: Session,
    *,
    state: str | None = None,
    region: str | None = None,
    min_pct_off: float | None = None,
    selected: str | None = None,
) -> list[str]:
    discovered = repo.list_distinct_categories(
        session, state=state, region=region, min_pct_off=min_pct_off
    )
    merged = list(discovered)
    if selected:
        cleaned = selected.strip()
        if cleaned and cleaned not in merged:
            merged.append(cleaned)
    merged = sorted(
        {*DEFAULT_CATEGORY_OPTIONS, *merged}, key=lambda value: value.casefold()
    )
    return merged


@app.get("/cheapskater")
def cheapskater_view(request: Request):
    """Render the Cheapskater deal board with the session-backed cart."""

    saved = _get_saved_deals(request)
    cart_items = list(saved.values())
    cart_total_items, cart_total_price = _cart_totals(saved)
    grouped = _group_saved_deals(saved)
    cart_groups = [
        {
            "store_number": store_number,
            "store_label": entries[0].get("store_label")
            or normalize_store_label(store_number, None),
            "deals": entries,
        }
        for store_number, entries in grouped.items()
    ]
    # Get user from session for nav
    user = None
    session_data = request.scope.get("session", {})
    if session_data.get("user_id"):
        from app.auth.dependencies import _get_db_session
        from app.auth.service import AuthService

        try:
            db = next(_get_db_session())
            user = AuthService(db).get_user_by_id(session_data["user_id"])
            db.close()
        except Exception:
            pass

    return templates.TemplateResponse(
        "cheapskater.html",
        {
            "request": request,
            "user": user,
            "cart_items": cart_items,
            "cart_groups": cart_groups,
            "cart_total_items": cart_total_items,
            "cart_total_price": cart_total_price,
            "format_currency": _format_currency,
            "active_scope": "cheapskater",
            "state": "ALL",
            "has_saved": bool(cart_items),
        },
    )


@app.post("/cheapskater/save-deal")
def save_deal(
    payload: SaveDealPayload,
    request: Request,
    session: Session = Depends(get_session),
) -> JSONResponse:
    """Persist or update a saved deal entry in the session cart."""

    store_number = (payload.store_number or "").strip()
    sku = (payload.sku or "").strip()
    if not store_number or not sku:
        raise HTTPException(
            status_code=400, detail="store_number and sku are required."
        )
    listing = _lookup_listing_for_save(session, store_number, sku)
    if listing is None:
        raise HTTPException(status_code=404, detail="Deal not found.")
    enriched = _prepare_listing(listing)
    deal = _build_cheapskater_deal(enriched)
    saved = _get_saved_deals(request)
    deal_id = deal.get("deal_id") or f"{store_number}:{sku}"
    quantity = int(payload.quantity or 1)
    entry = saved.get(deal_id)
    if entry:
        entry["quantity"] = int(entry.get("quantity", 0) or 0) + quantity
        entry["price"] = deal.get("price")
        entry["stock"] = deal.get("stock")
        entry["store_label"] = deal.get("store_label")
        entry["product_name"] = deal.get("product_name")
        entry["last_updated"] = deal.get("last_updated")
        entry["product_url"] = deal.get("product_url")
    else:
        entry = _saved_entry_from_deal(deal, quantity)
    saved[deal_id] = entry
    _persist_saved_deals(request, saved)
    cart_total_items, cart_total_price = _cart_totals(saved)
    return JSONResponse(
        content={
            "ok": True,
            "deal_id": deal_id,
            "cart_count": cart_total_items,
            "cart_total": cart_total_price,
            "deal": _serialize_saved_deal(entry),
        }
    )


@app.post("/cheapskater/cart/{deal_id}/increment")
def increment_saved_deal(deal_id: str, request: Request) -> JSONResponse:
    """Increase the planned quantity for a saved deal."""

    saved = _get_saved_deals(request)
    entry = saved.get(deal_id)
    if entry is None:
        raise HTTPException(status_code=404, detail="Deal not stored in cart.")
    entry["quantity"] = int(entry.get("quantity", 0) or 0) + 1
    saved[deal_id] = entry
    _persist_saved_deals(request, saved)
    cart_total_items, cart_total_price = _cart_totals(saved)
    return JSONResponse(
        content={
            "ok": True,
            "deal_id": deal_id,
            "quantity": entry["quantity"],
            "cart_count": cart_total_items,
            "cart_total": cart_total_price,
            "line_total": _deal_line_total(entry),
            "store_number": entry.get("store_number"),
            "store_label": entry.get("store_label"),
        }
    )


@app.post("/cheapskater/cart/{deal_id}/decrement")
def decrement_saved_deal(deal_id: str, request: Request) -> JSONResponse:
    """Decrease the planned quantity for a saved deal (removing when zero)."""

    saved = _get_saved_deals(request)
    entry = saved.get(deal_id)
    if entry is None:
        raise HTTPException(status_code=404, detail="Deal not stored in cart.")
    current_quantity = int(entry.get("quantity", 0) or 0) - 1
    deleted = current_quantity <= 0
    store_number = entry.get("store_number")
    store_label = entry.get("store_label")
    if deleted:
        saved.pop(deal_id, None)
    else:
        entry["quantity"] = current_quantity
        saved[deal_id] = entry
    _persist_saved_deals(request, saved)
    cart_total_items, cart_total_price = _cart_totals(saved)
    return JSONResponse(
        content={
            "ok": True,
            "deal_id": deal_id,
            "quantity": current_quantity if not deleted else 0,
            "cart_count": cart_total_items,
            "cart_total": cart_total_price,
            "line_total": _deal_line_total(entry) if not deleted else 0.0,
            "store_number": store_number,
            "store_label": store_label,
            "deleted": deleted,
        }
    )


def _render_dashboard(
    request: Request,
    *,
    scope: Scope,
    state: str | None,
    region: str | None,
    category: str | None,
    filters: dict[str, Any],
    session: Session,
    page_title: str | None = None,
    base_url: str = "/",
):
    LOGGER.debug(
        "Rendering dashboard",
        extra={"scope": scope, "state": state, "category": category, "region": region},
    )

    # Get user from session FIRST so we can use it for data filtering
    user = None
    session_data = request.scope.get("session", {})
    if session_data.get("user_id"):
        try:
            from app.auth.service import AuthService

            user = AuthService(session).get_user_by_id(session_data["user_id"])
            if user:
                LOGGER.debug(f"Loaded user {user.email} (ID: {user.id}) from session")
        except Exception as e:
            LOGGER.error(f"Failed to load user from session: {e}")
            pass

    is_pro, subscription_plan = _get_user_subscription_info(session, user)
    if user:
        LOGGER.debug(f"User {user.email} is_pro={is_pro}, plan={subscription_plan}")

    raw_items = _select_items(
        session,
        scope=scope,
        state=state if region != "FL" else None,
        category=category,
        region=region,
        user=user,
        is_paid=is_pro,
    )
    prepared_items = _prepare_listings(raw_items)
    # If explicit state was passed (which we might not want if region is hardcoded), force it
    # But wait, raw_items already filtered by region.
    # We might still want to filter by state WITHIN the region (e.g. WA vs OR).
    # For Florida, we don't need state filter if it's all FL.

    # Actually, _select_items handles the query.
    # _filter_by_state is a post-filter on the list.
    state_filtered = _filter_by_state(
        prepared_items, state if region == "WA_OR" else None
    )

    items = _apply_filters(state_filtered, filters=filters)
    # Count unique stores across all listings
    unique_store_ids = {item.get("store_id") for item in items if item.get("store_id")}
    store_count = len(unique_store_ids)
    grouped = _group_listings(items)
    grouped = _sort_groups(grouped, filters.get("sort_choice"))
    serialized_groups = [_serialize_group(group) for group in grouped]
    initial_groups = serialized_groups[:INITIAL_GROUP_BATCH]
    LOGGER.info(
        "Dashboard payload | groups=%s initial=%s stores=%s",
        len(serialized_groups),
        len(initial_groups),
        store_count,
    )
    categories = _collect_categories(
        session,
        state=state if region == "WA_OR" else None,
        region=region,
        min_pct_off=filters.get("discount_min"),
        selected=category,
    )
    last_updated = repo.get_latest_timestamp(session)

    # Default title
    if not page_title:
        page_title = "New Today" if scope == "new" else "Clearance Dashboard"
        if region == "WA_OR":
            page_title += " (Pacific NW)"

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "store_count": store_count,
            "groups": serialized_groups,
            "initial_groups": initial_groups,
            "group_json": json.dumps(serialized_groups, ensure_ascii=False),
            "initial_batch_size": INITIAL_GROUP_BATCH,
            "active_scope": scope,
            "state": state or "ALL",
            "region": region or "WA_OR",
            "current_region_slug": "florida" if region == "FL" else "wa_or",
            "page_title": page_title,
            "base_url": base_url,
            "category": category,
            "categories": categories,
            "state_options": STATE_OPTIONS if region == "WA_OR" else ["ALL", "FL"],
            "region_options": REGION_OPTIONS,
            "last_updated": last_updated,
            "state_from_zip": _state_from_zip,
            "store_url_builder": _store_specific_url,
            "store_label_builder": _format_store_label,
            "store_tooltip_builder": _format_store_tooltip,
            "format_timestamp": _format_timestamp,
            "format_currency": _format_currency,
            "format_regional_timestamp": format_regional_timestamp,
            "format_regional_full_timestamp": format_regional_full_timestamp,
            "filters": filters,
            "time_filter_options": TIME_FILTER_OPTIONS,
            "discount_filter_options": DISCOUNT_FILTER_OPTIONS,
            "stock_filter_options": STOCK_FILTER_OPTIONS,
            "sort_filter_options": SORT_OPTIONS,
            "is_pro": is_pro,
            "subscription_plan": subscription_plan,
        },
    )


@app.get("/healthz")
def healthcheck() -> dict[str, Any]:
    """Return health along with the age of the latest ZIP heartbeat."""

    payload = _read_json(ZIP_CURSOR_FILE)
    ts_text = payload.get("timestamp") if isinstance(payload, dict) else None
    if not ts_text:
        raise HTTPException(status_code=503, detail="zip cursor missing")
    try:
        stamp = datetime.fromisoformat(ts_text)
    except Exception as exc:
        raise HTTPException(
            status_code=503, detail=f"cursor parse error: {exc}"
        ) from exc
    age_minutes = (datetime.now(timezone.utc) - stamp).total_seconds() / 60.0
    if HEALTH_MAX_STALE_MINUTES > 0 and age_minutes > HEALTH_MAX_STALE_MINUTES:
        raise HTTPException(
            status_code=503,
            detail=f"stale cursor ({age_minutes:.1f}m > {HEALTH_MAX_STALE_MINUTES}m)",
        )
    return {"status": "ok", "age_minutes": age_minutes}


@app.get("/metrics")
def metrics() -> dict[str, Any]:
    """Return the latest aggregated metrics snapshot."""

    payload = _read_json(METRICS_SUMMARY_FILE)
    if not payload:
        raise HTTPException(status_code=404, detail="metrics summary unavailable")
    return payload


@app.get("/")
def list_clearance(
    request: Request,
    state: str | None = Query(None, description="State filter (WA or OR)."),
    category: str | None = Query(None, description="Optional category filter."),
    time_window: str = Query("all", description="Time filter window key."),
    discount_filter: str | None = Query(
        "50", description="Discount preset (percentage or custom)."
    ),
    discount_min: str | None = Query(
        None, description="Custom minimum discount percentage."
    ),
    discount_max: str | None = Query(
        None, description="Custom maximum discount percentage."
    ),
    stock_filter: str | None = Query(
        None, description="Stock preset (quantity or custom)."
    ),
    stock_min: str | None = Query(None, description="Custom minimum stock value."),
    stock_max: str | None = Query(None, description="Custom maximum stock value."),
    sort_order: str = Query("newest", description="Sort order key."),
    session: Session = Depends(get_session),
):
    """Render the Pacific NW (WA/OR) clearance dashboard."""

    normalized_state = _normalize_state(state)
    normalized_category = category or None
    filters = _normalize_filters(
        time_window=time_window,
        discount_filter=discount_filter,
        discount_min=discount_min,
        discount_max=discount_max,
        stock_filter=stock_filter,
        stock_min=stock_min,
        stock_max=stock_max,
        sort_order=sort_order,
    )

    return _render_dashboard(
        request,
        scope="all",
        state=normalized_state,
        region="WA_OR",
        category=normalized_category,
        filters=filters,
        session=session,
        base_url="",
    )


@app.get("/florida")
def list_florida_clearance(
    request: Request,
    category: str | None = Query(None, description="Optional category filter."),
    time_window: str = Query("all", description="Time filter window key."),
    discount_filter: str | None = Query(
        "50", description="Discount preset (percentage or custom)."
    ),
    discount_min: str | None = Query(
        None, description="Custom minimum discount percentage."
    ),
    discount_max: str | None = Query(
        None, description="Custom maximum discount percentage."
    ),
    stock_filter: str | None = Query(
        None, description="Stock preset (quantity or custom)."
    ),
    stock_min: str | None = Query(None, description="Custom minimum stock value."),
    stock_max: str | None = Query(None, description="Custom maximum stock value."),
    sort_order: str = Query("newest", description="Sort order key."),
    session: Session = Depends(get_session),
):
    """Render the Florida clearance dashboard."""

    normalized_category = category or None
    filters = _normalize_filters(
        time_window=time_window,
        discount_filter=discount_filter,
        discount_min=discount_min,
        discount_max=discount_max,
        stock_filter=stock_filter,
        stock_min=stock_min,
        stock_max=stock_max,
        sort_order=sort_order,
    )

    return _render_dashboard(
        request,
        scope="all",
        state="FL",
        region="FL",
        category=normalized_category,
        filters=filters,
        session=session,
        page_title="Florida Clearance",
        base_url="/florida",
    )


@app.get("/new-today")
def list_new_clearance_today(
    request: Request,
    state: str | None = Query(None, description="State filter (WA or OR)."),
    category: str | None = Query(None, description="Optional category filter."),
    time_window: str = Query("all", description="Time filter window key."),
    discount_filter: str | None = Query(
        "50", description="Discount preset (percentage or custom)."
    ),
    discount_min: str | None = Query(
        None, description="Custom minimum discount percentage."
    ),
    discount_max: str | None = Query(
        None, description="Custom maximum discount percentage."
    ),
    stock_filter: str | None = Query(
        None, description="Stock preset (quantity or custom)."
    ),
    stock_min: str | None = Query(None, description="Custom minimum stock value."),
    stock_max: str | None = Query(None, description="Custom maximum stock value."),
    sort_order: str = Query("newest", description="Sort order key."),
    session: Session = Depends(get_session),
):
    """Render new items for Pacific NW."""

    normalized_state = _normalize_state(state)
    normalized_category = category or None
    filters = _normalize_filters(
        time_window=time_window,
        discount_filter=discount_filter,
        discount_min=discount_min,
        discount_max=discount_max,
        stock_filter=stock_filter,
        stock_min=stock_min,
        stock_max=stock_max,
        sort_order=sort_order,
    )

    return _render_dashboard(
        request,
        scope="new",
        state=normalized_state,
        region="WA_OR",
        category=normalized_category,
        filters=filters,
        session=session,
        base_url="",
    )


@app.get("/florida/new-today")
def list_florida_new_clearance_today(
    request: Request,
    category: str | None = Query(None, description="Optional category filter."),
    time_window: str = Query("all", description="Time filter window key."),
    discount_filter: str | None = Query(
        "50", description="Discount preset (percentage or custom)."
    ),
    discount_min: str | None = Query(
        None, description="Custom minimum discount percentage."
    ),
    discount_max: str | None = Query(
        None, description="Custom maximum discount percentage."
    ),
    stock_filter: str | None = Query(
        None, description="Stock preset (quantity or custom)."
    ),
    stock_min: str | None = Query(None, description="Custom minimum stock value."),
    stock_max: str | None = Query(None, description="Custom maximum stock value."),
    sort_order: str = Query("newest", description="Sort order key."),
    session: Session = Depends(get_session),
):
    """Render new items for Florida."""

    normalized_category = category or None
    filters = _normalize_filters(
        time_window=time_window,
        discount_filter=discount_filter,
        discount_min=discount_min,
        discount_max=discount_max,
        stock_filter=stock_filter,
        stock_min=stock_min,
        stock_max=stock_max,
        sort_order=sort_order,
    )

    return _render_dashboard(
        request,
        scope="new",
        state="FL",
        region="FL",
        category=normalized_category,
        filters=filters,
        session=session,
        page_title="New Today (Florida)",
        base_url="/florida",
    )


@app.get("/export.xlsx")
def export_excel(
    request: Request,
    scope: Scope = Query("all", description="Dataset to export (all or new)."),
    state: str | None = Query(None, description="State filter (WA or OR)."),
    category: str | None = Query(None, description="Optional category filter."),
    time_window: str = Query("all", description="Time filter window key."),
    discount_filter: str | None = Query(
        "50", description="Discount preset (percentage or custom)."
    ),
    discount_min: str | None = Query(
        None, description="Custom minimum discount percentage."
    ),
    discount_max: str | None = Query(
        None, description="Custom maximum discount percentage."
    ),
    stock_filter: str | None = Query(
        None, description="Stock preset (quantity or custom)."
    ),
    stock_min: str | None = Query(None, description="Custom minimum stock value."),
    stock_max: str | None = Query(None, description="Custom maximum stock value."),
    sort_order: str = Query("newest", description="Sort order key."),
    session: Session = Depends(get_session),
) -> StreamingResponse:
    """Return an Excel workbook for the current filter selection."""
    # Requested to be removed/disabled to prevent abuse
    raise HTTPException(
        status_code=403, detail="Export to Excel is currently disabled."
    )

    # Original implementation commented out/unreachable below...
    # Get user for subscription check
    user = None
    session_data = request.scope.get("session", {})
    if session_data.get("user_id"):
        from app.auth.dependencies import _get_db_session
        from app.auth.service import AuthService

        try:
            db = next(_get_db_session())
            user = AuthService(db).get_user_by_id(session_data["user_id"])
            db.close()
        except Exception:
            pass

    normalized_state = _normalize_state(state)
    normalized_category = category or None
    filters = _normalize_filters(
        time_window=time_window,
        discount_filter=discount_filter,
        discount_min=discount_min,
        discount_max=discount_max,
        stock_filter=stock_filter,
        stock_min=stock_min,
        stock_max=stock_max,
        sort_order=sort_order,
    )
    raw_items = _select_items(
        session, scope=scope, state=None, category=normalized_category, user=user
    )
    prepared_items = _prepare_listings(raw_items)
    state_filtered = _filter_by_state(prepared_items, normalized_state)
    items = _apply_filters(state_filtered, filters=filters)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clearance"
    headers = [
        "Added (UTC)",
        "Last Updated (UTC)",
        "Price Change (UTC)",
        "State",
        "ZIP",
        "Store",
        "SKU",
        "Title",
        "Category",
        "Price",
        "Was",
        "% Off",
        "Availability",
        "Prev Price",
        "Prev Was",
        "Prev Updated (UTC)",
        "URL",
    ]
    sheet.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="004990", end_color="004990", fill_type="solid"
    )
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    sheet.freeze_panes = "A2"

    for item in items:
        pct_off = item.get("pct_off")
        pct_value = pct_off * 100 if pct_off is not None else None
        first_seen = item.get("first_seen") or item.get("price_started_at")
        last_seen = item.get("updated_at") or item.get("price_started_at")
        price_change = item.get("price_started_at")
        prev_updated = item.get("prev_updated_at")
        if isinstance(first_seen, datetime):
            first_seen = first_seen.isoformat()
        if isinstance(last_seen, datetime):
            last_seen = last_seen.isoformat()
        if isinstance(price_change, datetime):
            price_change = price_change.isoformat()
        if isinstance(prev_updated, datetime):
            prev_updated = prev_updated.isoformat()
        sheet.append(
            [
                first_seen,
                last_seen,
                price_change,
                _listing_state(item),
                item.get("store_zip"),
                item.get("store_name"),
                item.get("sku"),
                item.get("title"),
                item.get("category"),
                item.get("price"),
                item.get("price_was"),
                pct_value,
                item.get("availability"),
                item.get("prev_price"),
                item.get("prev_price_was"),
                prev_updated,
                item.get("product_url"),
            ]
        )

    for column_cells in sheet.iter_cols(
        min_row=1, max_row=sheet.max_row, max_col=sheet.max_column
    ):
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        if max_length <= 0:
            continue
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    state_label = normalized_state or "ALL"
    category_label = (normalized_category or "all").replace(" ", "-")
    filename = f"lowes-{scope}-{state_label.lower()}-{category_label.lower()}.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename={filename}",
    }
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/api/clearance")
def api_clearance(
    request: Request,
    scope: Scope = Query("all", description="Dataset to export (all or new)."),
    state: str | None = Query(None, description="State filter (WA or OR)."),
    category: str | None = Query(None, description="Optional category filter."),
    time_window: str = Query("all", description="Time filter window key."),
    discount_filter: str | None = Query(
        "50", description="Discount preset (percentage or custom)."
    ),
    discount_min: str | None = Query(
        None, description="Custom minimum discount percentage."
    ),
    discount_max: str | None = Query(
        None, description="Custom maximum discount percentage."
    ),
    stock_filter: str | None = Query(
        None, description="Stock preset (quantity or custom)."
    ),
    stock_min: str | None = Query(None, description="Custom minimum stock value."),
    stock_max: str | None = Query(None, description="Custom maximum stock value."),
    sort_order: str = Query("newest", description="Sort order key."),
    session: Session = Depends(get_session),
) -> JSONResponse:
    """Return clearance items as JSON data."""

    # Get user for subscription check
    user = None
    session_data = request.scope.get("session", {})
    if session_data.get("user_id"):
        from app.auth.dependencies import _get_db_session
        from app.auth.service import AuthService

        try:
            db = next(_get_db_session())
            user = AuthService(db).get_user_by_id(session_data["user_id"])
            db.close()
        except Exception:
            pass

    normalized_state = _normalize_state(state)
    normalized_category = category or None
    filters = _normalize_filters(
        time_window=time_window,
        discount_filter=discount_filter,
        discount_min=discount_min,
        discount_max=discount_max,
        stock_filter=stock_filter,
        stock_min=stock_min,
        stock_max=stock_max,
        sort_order=sort_order,
    )
    raw_items = _select_items(
        session,
        scope=scope,
        state=None,
        category=normalized_category,
        user=user,
    )
    prepared_items = _prepare_listings(raw_items)
    state_filtered = _filter_by_state(prepared_items, normalized_state)
    items = _apply_filters(state_filtered, filters=filters)
    grouped = _group_listings(items)
    grouped = _sort_groups(grouped, filters.get("sort_choice"))
    payload = [_serialize_observation(item) for item in items]
    grouped_payload = [_serialize_group(group) for group in grouped]
    filters_payload = {
        "time_window": filters.get("time_window"),
        "time_cutoff": filters.get("time_cutoff").isoformat()
        if filters.get("time_cutoff")
        else None,
        "discount_choice": filters.get("discount_choice"),
        "discount_min_pct": filters.get("discount_min_pct"),
        "discount_max_pct": filters.get("discount_max_pct"),
        "stock_choice": filters.get("stock_choice"),
        "stock_min": filters.get("stock_min"),
        "stock_max": filters.get("stock_max"),
        "sort_choice": filters.get("sort_choice"),
    }
    return JSONResponse(
        content={
            "items": payload,
            "groups": grouped_payload,
            "count": len(payload),
            "scope": scope,
            "state": normalized_state or "ALL",
            "category": normalized_category,
            "filters": filters_payload,
        }
    )


@app.get("/api/categories")
def api_categories(session: Session = Depends(get_session)) -> list[str]:
    """Return a sorted list of distinct clearance categories."""

    return repo.list_distinct_categories(session)


@app.get("/api/deals")
def api_deals(
    request: Request,
    category: str | None = Query(None, description="Optional category filter."),
    session: Session = Depends(get_session),
) -> JSONResponse:
    """Return clearance deals as a simple list (category-filterable)."""

    # Get user for subscription check
    user = None
    session_data = request.scope.get("session", {})
    if session_data.get("user_id"):
        from app.auth.dependencies import _get_db_session
        from app.auth.service import AuthService

        try:
            db = next(_get_db_session())
            user = AuthService(db).get_user_by_id(session_data["user_id"])
            db.close()
        except Exception:
            pass

    normalized_category = (category or "").strip() or None
    raw_items = _select_items(
        session, scope="all", state=None, category=normalized_category, user=user
    )
    prepared_items = _prepare_listings(raw_items)
    payload = [_serialize_observation(item) for item in prepared_items]
    return JSONResponse(
        content={
            "items": payload,
            "count": len(payload),
            "category": normalized_category,
        }
    )


@app.get("/api/stats")
def api_stats(session: Session = Depends(get_session)) -> JSONResponse:
    """Return aggregate stats for dashboard clients."""

    payload = _cache_stats(session)
    return JSONResponse(content=payload)


def _parse_store_info(store_id: str, store_name: str) -> dict[str, str]:
    """Parse store_name to extract city and state.

    Examples:
        "Seattle, WA (#0001)" -> city="Seattle", state="WA"
        "Portland #1234" -> city="Portland", state=""
    """
    city = ""
    state = ""
    match = re.match(r"^([^,]+),\s*([A-Z]{2})", store_name)
    if match:
        city = match.group(1).strip()
        state = match.group(2).strip()
    else:
        city = re.sub(r"\s*[#(].*", "", store_name).strip()
    return {"city": city, "state": state}


@app.post("/api/ingest")
def ingest_data(
    payload: IngestRequest,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    """Ingest deals from external workers (e.g., Gloorbot)."""

    if not payload.deals:
        return {"ok": True, "count": 0}

    upserted_count = 0
    skipped_count = 0

    known_stores = LOWES_STORES_WA_OR

    for deal in payload.deals:
        sku = _extract_sku(deal.product_url)
        if not sku:
            skipped_count += 1
            continue

        store_id = _normalize_store_number(deal.store_id)

        # Resolve store details
        store_details = _canonical_store_details(store_id)
        store_zip = "00000"
        store_city = None
        store_state = None
        store_region = None

        if store_details:
            store_zip = store_details.get("zip") or "00000"
            store_city = store_details.get("city")
            store_state = store_details.get("state")
            if store_state in ("WA", "OR"):
                store_region = "WA_OR"
        else:
            # Fallback: Parse from Gloorbot provided name (critical for FL stores)
            parsed = _parse_store_info(store_id, deal.store_name)
            store_city = parsed["city"]
            store_state = parsed["state"]

            if store_state == "FL":
                store_region = "FL"
            elif store_state in ("WA", "OR"):
                store_region = "WA_OR"

        # Ensure Store exists
        repo.upsert_store(
            session,
            store_id=store_id,
            name=deal.store_name,
            zip_code=store_zip,
            city=store_city,
            state=store_state,
            region=store_region,
        )

        preferred_category = (deal.category_name or "").strip()
        category_name = preferred_category or _extract_category_name(deal.category_url)

        image_url = deal.image_url
        image_url = _sanitize_image_url(image_url)

        # Ensure Item exists
        repo.upsert_item(
            session,
            sku=sku,
            retailer="lowes",
            title=deal.title,
            category=category_name,
            product_url=deal.product_url,
            image_url=image_url,
        )

        ts = datetime.fromisoformat(deal.found_at.replace("Z", "+00:00"))

        # Add Observation
        observation = Observation(
            ts_utc=ts,
            retailer="lowes",
            store_id=store_id,
            store_name=deal.store_name,
            zip=store_zip,
            sku=sku,
            title=deal.title,
            category=category_name,
            price=deal.price,
            price_was=deal.was_price,
            pct_off=deal.pct_off,
            availability=None,  # Gloorbot doesn't send availability yet
            product_url=deal.product_url,
            image_url=image_url,
            clearance=True,  # Assuming pushed deals are clearance
        )
        repo.insert_observation(session, observation)

        # Update History
        repo.update_price_history(
            session,
            retailer="lowes",
            store_id=store_id,
            sku=sku,
            title=deal.title,
            category=category_name,
            ts_utc=ts,
            price=deal.price,
            price_was=deal.was_price,
            pct_off=deal.pct_off,
            availability=None,
            product_url=deal.product_url,
            image_url=image_url,
            clearance=True,
        )
        upserted_count += 1

    session.commit()
    LOGGER.info(
        f"Ingested {upserted_count} deals from {payload.source} (skipped {skipped_count})"
    )
    return {"ok": True, "count": upserted_count, "skipped": skipped_count}


@app.post("/api/setup-admin")
def setup_admin_account(session: Session = Depends(get_session)) -> JSONResponse:
    """
    ONE-TIME ENDPOINT: Create admin account on production.
    This endpoint should be removed after use for security.
    """
    from app.auth.models import User, Subscription, SubscriptionPlan, SubscriptionStatus
    from app.auth.service import AuthService

    try:
        auth_service = AuthService(session)

        # Check if user exists
        existing_user = (
            session.query(User).filter(User.email == "93robingattis@gmail.com").first()
        )

        if existing_user:
            user = existing_user
            message = f"User already exists with ID: {user.id}"
        else:
            # Create user
            user = auth_service.register_user(
                email="93robingattis@gmail.com",
                password="Alphonse5150$",
                display_name="Robin Gattis",
            )
            message = f"Created new user with ID: {user.id}"

        # Check if subscription exists
        existing_sub = (
            session.query(Subscription).filter(Subscription.user_id == user.id).first()
        )

        if existing_sub:
            # Update to PRO
            existing_sub.plan = SubscriptionPlan.PRO
            existing_sub.status = SubscriptionStatus.ACTIVE
            existing_sub.current_period_end = datetime.now(timezone.utc) + timedelta(
                days=365
            )
            message += " | Updated subscription to PRO"

        session.commit()

        return JSONResponse(
            content={
                "ok": True,
                "message": message,
                "email": "93robingattis@gmail.com",
                "plan": "PRO",
                "user_id": user.id,
            }
        )

    except Exception as e:
        session.rollback()
        LOGGER.exception("Failed to create admin account")
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn

    port = int(os.getenv("PORT", "8000"))
    uvicorn.run("app.dashboard:app", host="0.0.0.0", port=port, reload=False)
